import hashlib
import csv
import json
import math
import os
import random
from copy import deepcopy

from openpyxl import Workbook, load_workbook
import xlrd
import yaml

# 全局表格缓存（避免重复读取文件）
_DF_CACHE = {}


class DataTable:
    """轻量二维表结构，用于替代 pandas DataFrame 的核心能力。"""

    def __init__(self, rows=None, columns=None):
        self.rows = rows or []
        self.columns = columns or self._infer_columns(self.rows)

    @staticmethod
    def _infer_columns(rows):
        cols = []
        seen = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    cols.append(key)
        return cols

    def __len__(self):
        return len(self.rows)

    def copy(self):
        return DataTable(rows=deepcopy(self.rows), columns=list(self.columns))

    def with_row_id(self, column_name="row_id"):
        new_rows = []
        for idx, row in enumerate(self.rows):
            item = dict(row)
            item[column_name] = f"row_{idx}"
            new_rows.append(item)
        new_cols = list(self.columns)
        if column_name not in new_cols:
            new_cols.append(column_name)
        return DataTable(rows=new_rows, columns=new_cols)

    def shuffled(self):
        new_rows = deepcopy(self.rows)
        random.shuffle(new_rows)
        return DataTable(rows=new_rows, columns=list(self.columns))

    def filter_not_in(self, column_name, values):
        value_set = set(str(v) for v in values)
        kept = []
        for row in self.rows:
            current = str(row.get(column_name, ""))
            if current not in value_set:
                kept.append(row)
        return DataTable(rows=kept, columns=list(self.columns))

    def to_records(self):
        return deepcopy(self.rows)


def is_null_value(value):
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def _read_csv_rows(file_path):
    try:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = [dict(row) for row in reader]
            columns = list(reader.fieldnames or [])
            return DataTable(rows=rows, columns=columns)
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="gbk", newline="") as f:
            reader = csv.DictReader(f)
            rows = [dict(row) for row in reader]
            columns = list(reader.fieldnames or [])
            return DataTable(rows=rows, columns=columns)


def _read_xlsx_rows(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return DataTable(rows=[], columns=[])

    header = [str(h) if h is not None else "" for h in all_rows[0]]
    columns = [c if c else f"col_{idx}" for idx, c in enumerate(header)]
    rows = []
    for row_values in all_rows[1:]:
        row_dict = {}
        for idx, col_name in enumerate(columns):
            row_dict[col_name] = row_values[idx] if idx < len(row_values) else None
        rows.append(row_dict)
    return DataTable(rows=rows, columns=columns)


def _read_xls_rows(file_path):
    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_index(0)
    if sheet.nrows == 0:
        return DataTable(rows=[], columns=[])

    header = [str(sheet.cell_value(0, c)) if sheet.cell_value(0, c) is not None else "" for c in range(sheet.ncols)]
    columns = [c if c else f"col_{idx}" for idx, c in enumerate(header)]
    rows = []
    for r in range(1, sheet.nrows):
        row_dict = {}
        for c, col_name in enumerate(columns):
            row_dict[col_name] = sheet.cell_value(r, c)
        rows.append(row_dict)
    return DataTable(rows=rows, columns=columns)


def _read_jsonl_rows(file_path):
    rows = []
    columns = []
    seen = set()
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            item = json.loads(line)
            if isinstance(item, dict):
                rows.append(item)
                for k in item.keys():
                    if k not in seen:
                        seen.add(k)
                        columns.append(k)
    return DataTable(rows=rows, columns=columns)


def _collect_columns(records):
    columns = []
    seen = set()
    for record in records:
        for key in record.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def write_records_csv(records, output_path):
    columns = _collect_columns(records)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for record in records:
            writer.writerow({col: record.get(col, "") for col in columns})


def write_records_jsonl(records, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def write_records_excel(records, output_path):
    columns = _collect_columns(records)
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for record in records:
        ws.append([record.get(col, "") for col in columns])
    wb.save(output_path)


def universal_read_data(file_path):
    """
    通用数据读取（返回 DataTable）
    带缓存机制，避免重复读取相同文件。
    """
    if not file_path or not os.path.exists(file_path):
        return None
    
    ext = os.path.splitext(file_path)[1].lower()
    
    # 检查缓存
    try:
        mtime = os.path.getmtime(file_path)
        cache_key = file_path
        
        if cache_key in _DF_CACHE:
            cached_mtime, cached_table = _DF_CACHE[cache_key]
            if cached_mtime == mtime:
                # 缓存命中！返回副本（防止修改原数据）
                return cached_table.copy()
    except Exception:
        pass  # 缓存失败，继续读取
    
    # 缓存未命中，读取文件
    if ext == ".csv":
        table = _read_csv_rows(file_path)
    elif ext == ".xlsx":
        table = _read_xlsx_rows(file_path)
    elif ext == ".xls":
        table = _read_xls_rows(file_path)
    elif ext == ".jsonl":
        table = _read_jsonl_rows(file_path)
    else:
        raise ValueError(f"不支持的数据格式: {ext}")

    # 存入缓存
    try:
        mtime = os.path.getmtime(file_path)
        _DF_CACHE[file_path] = (mtime, table.copy())
    except Exception:
        pass  # 缓存失败不影响数据读取

    return table


def get_task_hash(yaml_str, file_path):
    try:
        config = yaml.safe_load(yaml_str)
        task_cfg_str = json.dumps(config.get("task", {}), sort_keys=True)
    except Exception:
        task_cfg_str = yaml_str

    hasher = hashlib.md5()
    hasher.update(task_cfg_str.encode("utf-8"))

    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096 * 1024), b""):
                hasher.update(chunk)
    else:
        hasher.update(file_path.encode("utf-8"))

    return hasher.hexdigest()
