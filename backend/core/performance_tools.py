"""
性能优化工具类

提供防抖、节流、异步加载等工具
"""

import logging
import threading
from typing import Callable, Any
from functools import wraps
from PySide6.QtCore import QTimer, Signal, QObject
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class Debouncer:
    """
    防抖器 - 在指定时间内多次触发时，只执行最后一次
    
    典型用途：搜索输入、配置变化监听
    """

    def __init__(self, wait_ms: int = 500):
        """
        Args:
            wait_ms: 防抖延迟（毫秒）
        """
        self.wait_ms = wait_ms
        self.timer = None
        self.callbacks = []

    def add_callback(self, callback: Callable):
        """添加回调函数"""
        self.callbacks.append(callback)
        return self

    def call(self, *args, **kwargs):
        """触发防抖"""
        # 取消之前的定时器
        if self.timer:
            self.timer.stop()

        # 创建新定时器
        self.timer = QTimer()
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(lambda: self._execute(*args, **kwargs))
        self.timer.start(self.wait_ms)

    def _execute(self, *args, **kwargs):
        """执行所有回调"""
        for callback in self.callbacks:
            try:
                callback(*args, **kwargs)
            except Exception as e:
                logger.error(f"防抖回调执行失败: {e}")

    def cancel(self):
        """取消待处理的调用"""
        if self.timer:
            self.timer.stop()


class AsyncTaskRunner(QObject):
    """
    异步任务运行器 - 在后台线程执行耗时操作
    
    信号:
        task_finished: 任务完成，携带结果
        task_failed: 任务失败，携带异常
    """

    task_finished = Signal(object)
    task_failed = Signal(str)

    def __init__(self, max_workers: int = 2):
        super().__init__()
        self.executor = ThreadPoolExecutor(max_workers=max_workers)
        self.logger = logging.getLogger("AsyncTaskRunner")

    def execute_async(self, task_func: Callable, *args, **kwargs):
        """
        异步执行任务

        Args:
            task_func: 要执行的函数
            *args: 位置参数
            **kwargs: 关键字参数

        Returns:
            Future: 任务执行的 Future 对象
        """

        def wrapper():
            try:
                result = task_func(*args, **kwargs)
                self.task_finished.emit(result)
            except Exception as e:
                self.logger.error(f"异步任务失败: {e}", exc_info=True)
                self.task_failed.emit(str(e))

        future = self.executor.submit(wrapper)
        return future

    def shutdown(self):
        """关闭线程池"""
        self.executor.shutdown(wait=True)


class ProgressiveFileReader:
    """
    渐进式文件读取器 - 分块读取大文件，避免内存爆炸
    """

    @staticmethod
    def read_large_file(file_path: str, chunk_size: int = 10000):
        """
        分块读取文件

        Args:
            file_path: 文件路径
            chunk_size: 每块行数

        Yields:
            list[dict] 块
        """
        import csv
        import json
        try:
            if file_path.endswith(".csv"):
                with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f)
                    chunk = []
                    for row in reader:
                        chunk.append(dict(row))
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                    if chunk:
                        yield chunk
                return

            if file_path.endswith(".jsonl"):
                with open(file_path, "r", encoding="utf-8") as f:
                    chunk = []
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        item = json.loads(line)
                        if isinstance(item, dict):
                            chunk.append(item)
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                    if chunk:
                        yield chunk
                return

            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            iter_rows = ws.iter_rows(values_only=True)
            header = next(iter_rows, None)
            if not header:
                return

            columns = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(header)]
            chunk = []
            for raw_row in iter_rows:
                row = {columns[i]: (raw_row[i] if i < len(raw_row) else None) for i in range(len(columns))}
                chunk.append(row)
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []
            if chunk:
                yield chunk

        except Exception as e:
            logger.error(f"读取文件失败: {e}")
            raise


class TableOptimizer:
    """
    表格优化工具 - 快速填充表格数据
    """

    @staticmethod
    def fast_populate_table(table_widget, data_frame):
        """
        快速填充表格

        Args:
            table_widget: QTableWidget
            data_frame: list[dict] 或兼容 DataFrame 风格对象
        """
        from PySide6.QtWidgets import QTableWidgetItem

        columns = []
        rows = []

        if data_frame is None:
            table_widget.setRowCount(0)
            table_widget.setColumnCount(0)
            return

        # DataTable / DataFrame-like
        if hasattr(data_frame, "columns") and hasattr(data_frame, "rows"):
            columns = [str(col) for col in data_frame.columns]
            rows = data_frame.rows
        elif isinstance(data_frame, list):
            rows = data_frame
            seen = set()
            for row in rows:
                if isinstance(row, dict):
                    for key in row.keys():
                        key_text = str(key)
                        if key_text not in seen:
                            seen.add(key_text)
                            columns.append(key_text)
        elif hasattr(data_frame, "to_dict") and hasattr(data_frame, "columns"):
            # 兼容尚未迁移的 DataFrame
            columns = [str(col) for col in data_frame.columns]
            rows = data_frame.to_dict(orient="records")

        table_widget.setRowCount(len(rows))
        table_widget.setColumnCount(len(columns))
        table_widget.setHorizontalHeaderLabels(columns)

        for row_idx, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            for col_idx, col in enumerate(columns):
                try:
                    item = QTableWidgetItem(str(row.get(col, "")))
                    table_widget.setItem(row_idx, col_idx, item)
                except Exception:
                    pass

        # 最后调整列宽
        table_widget.resizeColumnsToContents()


def performance_timer(func):
    """
    性能计时装饰器 - 记录函数执行时间
    
    使用:
        @performance_timer
        def my_slow_function():
            ...
    """

    @wraps(func)
    def wrapper(*args, **kwargs):
        import time

        start = time.time()
        try:
            result = func(*args, **kwargs)
            duration = time.time() - start
            if duration > 0.1:  # 仅记录超过100ms的调用
                logger.warning(f"[性能] {func.__name__} 耗时 {duration*1000:.1f}ms")
            return result
        except Exception as e:
            duration = time.time() - start
            logger.error(f"[性能] {func.__name__} 异常耗时 {duration*1000:.1f}ms: {e}")
            raise

    return wrapper
